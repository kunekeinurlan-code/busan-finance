from flask import Flask, request, jsonify, send_file, session
import os, json, io, math
from datetime import datetime
from collections import defaultdict
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "busan-finance-2026")

# Пароль — задаётся через переменную окружения PASSWORD на Railway
ACCESS_PASSWORD = os.environ.get("PASSWORD", "busan2026")

# Накопительная база — хранится в /tmp/busan_db.json
DB_PATH = "/tmp/busan_db.json"

NAVY="1A1A2E"; GOLD="D4A017"; WHITE="FFFFFF"; LIGHT="F5F5F0"
GREEN="27AE60"; RED="C0392B"; GRAY="CCCCCC"

def fill(c): return PatternFill("solid", fgColor=c)
def font(bold=False, color=WHITE, size=11, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def border_thin():
    s = Side(style="thin", color=GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

CATEGORIES = {
    "Перевод собственнику":  ["карту kaspi gold","kaspi gold *"],
    "Выручка Kaspi.kz":      ["продажи с kaspi.kz","возмещение коммерсанту"],
    "Зарплата / ФОТ":        ["оплата заработной платы","заработная плата","зарплата","salary","оклад","выплата зарплат","фот"],
    "Налоги / взносы":       ["налог","ипн","соц","пенсион","опв","снп","осмс","источника выплаты","облагаемых","единый совокупный","астана ерц","ерц"],
    "Аренда":                ["аренд","rent","найм"],
    "Реклама / маркетинг":   ["реклам","маркетинг","продвижен","smm","таргет","instagram","facebook","google ads","яндекс директ","объявлен","рекламных услуг"],
    "Процессинг / эквайринг":["процессинг","эквайринг","расчеты по картам","расчеты по карточкам","pos-терминал","pos терминал","acquiring","расчёты по карт","услуги процессинга","услуг процессинга","оплата за услуги операций по картам"],
    "Логистика / курьер":    ["доставк","логистик","курьер","express","dhl","cdek","сдэк","почтовые"],
    "Кредит / займ":         ["погашение","бизнес кредит","кредит","займ","ссуда","резервирование средств"],
    "Возврат покупателю":    ["возврат продаж","возврат средств","возврат оплаты","за непредоставленные усл"],
    "Продукты / питание":    ["магазин","супермарк","продукт","еда","food","market","magnum","small"],
    "Кафе / рестораны":      ["кафе","ресторан","cafe","restaurant","coffee","кофе"],
    "Транспорт":             ["такси","uber","yandex go","яндекс такси","bolt","автобус","транспорт","parking"],
    "Связь / интернет":      ["beeline","kcell","activ","altel","tele2","билайн","интернет","связь"],
    "Подписки / сервисы":    ["netflix","spotify","apple","youtube","подписк"],
    "Образование":           ["курс","обучен","образован","школ","универс","тренинг","семинар"],
    "Здоровье / медицина":   ["аптек","клиник","больниц","медицин","pharmacy","dental","стомат","europharma"],
    "Красота / уход":        ["салон","beauty","spa","cosmetic","косметик"],
    "Банковские расходы":    ["комисси","обслуживан","штраф","пени","оплата за информационно","оплата услуги по обработке","оплата услуг по обработке","за операций по картам","страховую премию","страхован"],
    "IT / сервисы":          ["it","разработк","программ","software","хостинг","домен"],
    "Инвестиции":            ["инвестиц","депозит","брокер","акци","облигац"],
    "Профессиональные услуги":["профессиональные","научные и технические","консультац","юридич","бухгалтер","нотариус"],
    "Доход от клиентов":     ["поступление от клиент","оплата от клиент","гонорар","вознагражд","выручка","комисс омк","перевод денежных средств по дог"],
    "Выручка по карточкам":  ["расчеты по карточкам","расчёты по карточкам","расчеты по картам","расчёты по картам","зачисление по pos","зачисл по pos","by card","эквайринг зачислен"],
    "Торговая выручка":      ["за товары","оплата за товар","реализац товар","расчет за товар"],
}


# Признаки внутренних переводов (между своими счетами) — исключаются из P&L
INTERNAL_PATTERNS = [
    "kaspipay на депозит",
    "депозит u35337359",
    "перевод со счета u35337359",
    "перевод со счета kaspipay",
    "со своего kaspi gold на счет в kaspi pay",   # поступление от собственника — внутреннее
    "собственных средств на свой счет в другом банке",
    "переводы между счетами",
    "переводы клиентом денег с одного своего",
    "без ндс. переводы клиентом",
    "между своими счетами",
    "между счетами. без ндс",
    "перевод с карты на счет ип",
]

def is_internal(desc):
    d = str(desc).lower()
    return any(p in d for p in INTERNAL_PATTERNS)

def categorize(desc):
    d = str(desc).lower()
    for cat, kws in CATEGORIES.items():
        if any(k in d for k in kws): return cat
    return "Прочие"

def parse_amount(val):
    if val is None: return 0.0
    import math
    if isinstance(val, float) and math.isnan(val): return 0.0
    try:
        s = str(val).strip()
        # Убираем все виды пробелов и неразрывные пробелы
        s = s.replace("\xa0","").replace("\u202f","").replace("\u00a0","").replace(" ","")
        s = s.replace(",",".")
        return float(s)
    except: return 0.0

def clean(v):
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)): return 0.0
    return v

def detect_bank(filename, filepath):
    fn = filename.lower()
    if "kaspi" in fn: return "Kaspi"
    if "бцк" in fn or "bcc" in fn or "bck" in fn or "centercredit" in fn or "цк" in fn: return "БЦК"
    if "halyk" in fn or "народн" in fn or "hsbk" in fn: return "Halyk"
    if "freedom" in fn or "ffin" in fn: return "Freedom"
    try:
        df = pd.read_excel(filepath, header=None, nrows=10)
        text = " ".join(df.fillna("").astype(str).values.flatten()).lower()
        if "kaspi" in text: return "Kaspi"
        if "центркредит" in text or "centercredit" in text or "kcjbkzkx" in text: return "БЦК"
        if "halyk" in text or "народный" in text or "hsbkkzkx" in text or "народный банк" in text: return "Halyk"
        if "freedom" in text or "ffin" in text: return "Freedom"
    except: pass
    return "Другой"



def parse_kaspi(filepath):
    """Парсер выписки Kaspi Bank"""
    transactions = []
    try:
        engine = "xlrd" if filepath.endswith(".xls") else "openpyxl"
        raw = pd.read_excel(filepath, header=None, engine=engine)

        header_row = None
        for i, row in raw.iterrows():
            vals = [str(v) for v in row.values]
            if any("Дата операции" in v or "Дата" in v for v in vals) and any("Дебет" in v for v in vals):
                header_row = i
                break
        if header_row is None:
            return transactions

        df = pd.read_excel(filepath, skiprows=header_row, header=0, engine=engine)
        df.columns = [str(c).strip() for c in df.columns]

        date_col   = next((c for c in df.columns if "Дата" in c), None)
        debit_col  = next((c for c in df.columns if c.strip() == "Дебет"), None)
        credit_col = next((c for c in df.columns if c.strip() == "Кредит"), None)
        desc_col   = next((c for c in df.columns if "Назначение" in c), None)
        name_col   = next((c for c in df.columns if "бенефициара" in c or "Наименование" in c), None)

        if not date_col: return transactions

        for _, row in df.iterrows():
            try:
                date_str = str(row.get(date_col, "")).strip()
                if not date_str or date_str == "nan": continue
                date = pd.to_datetime(date_str[:10], dayfirst=True, errors="coerce")
                if pd.isna(date): continue

                debit  = parse_amount(row.get(debit_col,  0)) if debit_col  else 0.0
                credit = parse_amount(row.get(credit_col, 0)) if credit_col else 0.0
                if debit == 0 and credit == 0: continue

                desc = str(row.get(desc_col, "")).strip() if desc_col else ""
                if desc == "nan" or not desc:
                    desc = str(row.get(name_col, "")).strip() if name_col else ""
                if desc == "nan": desc = ""

                cat = categorize(desc)
                internal = is_internal(desc)
                transactions.append({
                    "date":  date.strftime("%Y-%m-%d"),
                    "month": date.strftime("%Y-%m"),
                    "bank":  "Kaspi",
                    "description": desc,
                    "debit":  debit,
                    "credit": credit,
                    "category": "Внутренний перевод" if internal else cat,
                    "type":   "expense" if debit > 0 else "income",
                    "amount": debit if debit > 0 else credit,
                    "internal": internal,
                })
            except: continue
    except Exception as e:
        print(f"Kaspi parse error: {e}")
    return transactions


def parse_halyk(filepath):
    """Парсер выписки Halyk Bank (Народный Банк)"""
    transactions = []
    try:
        engine = "xlrd" if filepath.endswith(".xls") else "openpyxl"
        raw = pd.read_excel(filepath, header=None, engine=engine)

        # Ищем строку заголовка — содержит Дебет И Кредит И Назначение
        header_row = None
        for i, row in raw.iterrows():
            vals = [str(v) for v in row.values]
            has_debit   = any("Дебет" in v for v in vals)
            has_credit  = any("Кредит" in v for v in vals)
            has_purpose = any("Назначение" in v or "назначение" in v for v in vals)
            if has_debit and has_credit and has_purpose:
                header_row = i
                break

        if header_row is None:
            return transactions

        df = pd.read_excel(filepath, skiprows=header_row, header=0, engine=engine)
        df.columns = [str(c).strip() for c in df.columns]

        # Ищем колонки по названию
        date_col    = next((c for c in df.columns if "Дата" in c), None)
        debit_col   = next((c for c in df.columns if c.strip() == "Дебет"), None)
        credit_col  = next((c for c in df.columns if c.strip() == "Кредит"), None)
        desc_col    = next((c for c in df.columns if "Назначение" in c), None)
        # Fallback — берём последнюю текстовую колонку перед числовыми
        if not desc_col:
            desc_col = next((c for c in reversed(df.columns) if "назначен" in c.lower() or "платеж" in c.lower()), None)

        if not date_col or not desc_col:
            return transactions

        for _, row in df.iterrows():
            try:
                date_str = str(row.get(date_col, "")).strip()
                if not date_str or date_str in ("nan", "Итого оборот", "Исходящий"):
                    continue
                date = pd.to_datetime(date_str[:10], dayfirst=True, errors="coerce")
                if pd.isna(date):
                    continue

                debit  = parse_amount(row.get(debit_col,  0)) if debit_col  else 0.0
                credit = parse_amount(row.get(credit_col, 0)) if credit_col else 0.0
                if debit == 0 and credit == 0:
                    continue

                desc = str(row.get(desc_col, "")).strip()
                if desc in ("nan", ""):
                    # Попробуем взять из колонки Контрагент
                    cont_col = next((c for c in df.columns if "Контрагент" in c), None)
                    if cont_col:
                        desc = str(row.get(cont_col, "")).strip()
                if desc == "nan":
                    desc = ""

                cat = categorize(desc)
                internal = is_internal(desc)

                # Расчеты по карточкам Halyk — это выручка
                if "расчеты по карточкам" in desc.lower() or "расчёты по карточкам" in desc.lower():
                    cat = "Выручка по карточкам"
                    internal = False

                transactions.append({
                    "date":  date.strftime("%Y-%m-%d"),
                    "month": date.strftime("%Y-%m"),
                    "bank":  "Halyk",
                    "description": desc,
                    "debit":  debit,
                    "credit": credit,
                    "category": "Внутренний перевод" if internal else cat,
                    "type":   "expense" if debit > 0 else "income",
                    "amount": debit if debit > 0 else credit,
                    "internal": internal,
                })
            except:
                continue
    except Exception as e:
        print(f"Halyk parse error: {e}")
    return transactions

def parse_bcc(filepath):
    """Парсер выписки Банк ЦентрКредит (БЦК)"""
    transactions = []
    try:
        engine = "xlrd" if filepath.endswith(".xls") else "openpyxl"
        raw = pd.read_excel(filepath, header=None, engine=engine)

        # Ищем строку-заголовок (содержит "Дебет" и "Кредит")
        header_row = None
        for i, row in raw.iterrows():
            vals = [str(v) for v in row.values]
            if any("Дебет" in v or "Дата" in v for v in vals) and any("Кредит" in v for v in vals):
                header_row = i
                break
        if header_row is None:
            return transactions

        df = pd.read_excel(filepath, skiprows=header_row, header=0, engine=engine)
        df.columns = [str(c).strip() for c in df.columns]

        # Находим нужные колонки
        date_col  = next((c for c in df.columns if "Дата" in c or "Күні" in c), None)
        debit_col = next((c for c in df.columns if "Дебет" in c), None)
        credit_col= next((c for c in df.columns if "Кредит" in c and "конверт" not in c.lower()), None)
        desc_col  = next((c for c in df.columns if "Назначение" in c or "мақсаты" in c), None)

        if not date_col: return transactions

        for _, row in df.iterrows():
            try:
                date_str = str(row.get(date_col, "")).strip()
                if not date_str or date_str == "nan": continue
                # Формат даты БЦК: "26.12.2025 06:10:02"
                date = pd.to_datetime(date_str[:10], dayfirst=True, errors="coerce")
                if pd.isna(date): continue

                debit  = parse_amount(row.get(debit_col,  0)) if debit_col  else 0.0
                credit = parse_amount(row.get(credit_col, 0)) if credit_col else 0.0

                if debit == 0 and credit == 0: continue

                desc = str(row.get(desc_col, "")).strip() if desc_col else ""
                if desc == "nan": desc = ""

                cat = categorize(desc)
                internal = is_internal(desc)
                transactions.append({
                    "date":  date.strftime("%Y-%m-%d"),
                    "month": date.strftime("%Y-%m"),
                    "bank":  "БЦК",
                    "description": desc,
                    "debit":  debit,
                    "credit": credit,
                    "category": "Внутренний перевод" if internal else cat,
                    "type":   "expense" if debit > 0 else "income",
                    "amount": debit if debit > 0 else credit,
                    "internal": internal,
                })
            except: continue
    except Exception as e:
        print(f"БЦК parse error: {e}")
    return transactions


def parse_file(filepath, bank):
    transactions = []
    try:
        # БЦК имеет нестандартный формат — специальный парсер
        if bank == "БЦК":
            return parse_bcc(filepath)
        if bank == "Kaspi":
            return parse_kaspi(filepath)
        if bank in ("Halyk", "Народный"):
            return parse_halyk(filepath)

        raw = pd.read_excel(filepath, header=None, nrows=20)
        header_row = None
        for i, row in raw.iterrows():
            vals = [str(v).lower() for v in row.values]
            score = sum(1 for v in vals if any(k in v for k in ["дата","date","сумм","дебет","кредит","списан","зачислен"]))
            if score >= 2: header_row = i; break
        if header_row is None: header_row = 0
        df = pd.read_excel(filepath, skiprows=header_row, header=0)
        df.columns = [str(c).strip() for c in df.columns]
        date_col = next((c for c in df.columns if any(k in c.lower() for k in ["дата","date"])), None)
        if not date_col: return []
        for _, row in df.iterrows():
            try:
                date = pd.to_datetime(str(row.get(date_col,"")), dayfirst=True, errors="coerce")
                if pd.isna(date): continue
                debit = credit = 0.0
                for col in df.columns:
                    v = parse_amount(row.get(col, 0))
                    cl = col.lower()
                    if any(k in cl for k in ["списан","расход","дебет","debit","дт"]): debit = abs(v)
                    elif any(k in cl for k in ["зачислен","приход","кредит","credit","кт"]): credit = abs(v)
                if debit == 0 and credit == 0: continue
                desc = ""
                for col in df.columns:
                    if any(k in col.lower() for k in ["назначен","описан","получатель","контрагент","purpose"]):
                        v = str(row.get(col,"")).strip()
                        if v and v.lower() != "nan": desc = v; break
                cat = categorize(desc)
                internal = is_internal(desc)
                transactions.append({
                    "date": date.strftime("%Y-%m-%d"),
                    "month": date.strftime("%Y-%m"),
                    "bank": bank, "description": desc,
                    "debit": debit, "credit": credit,
                    "category": "Внутренний перевод" if internal else cat,
                    "type": "expense" if debit > 0 else "income",
                    "amount": debit if debit > 0 else credit,
                    "internal": internal,
                })
            except: continue
    except Exception as e:
        print(f"Parse error: {e}")
    return transactions

# ─── НАКОПИТЕЛЬНАЯ БАЗА ───────────────────────────────────────────────────────

def load_db():
    try:
        if os.path.exists(DB_PATH):
            with open(DB_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
    except: pass
    return {"transactions": [], "updated_at": None}

def save_db(db):
    with open(DB_PATH, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False)

def merge_transactions(existing, new_txns):
    """Добавляет только новые транзакции (дедупликация по дате+сумме+описанию)"""
    existing_keys = set(
        f"{t['date']}_{t['debit']}_{t['credit']}_{t['description'][:30]}"
        for t in existing
    )
    added = 0
    for t in new_txns:
        key = f"{t['date']}_{t['debit']}_{t['credit']}_{t['description'][:30]}"
        if key not in existing_keys:
            existing.append(t)
            existing_keys.add(key)
            added += 1
    return added

# ─── АНАЛИТИКА ────────────────────────────────────────────────────────────────

def build_analytics(all_transactions):
    # Внутренние переводы между своими счетами исключаем из P&L
    business_txns = [t for t in all_transactions if not t.get("internal")]
    internal_txns = [t for t in all_transactions if t.get("internal")]

    total_inc = clean(sum(t["credit"] for t in business_txns))
    total_exp = clean(sum(t["debit"]  for t in business_txns))
    internal_out = clean(sum(t["debit"]  for t in internal_txns))
    internal_in  = clean(sum(t["credit"] for t in internal_txns))
    net = clean(total_inc - total_exp)

    cat_exp = defaultdict(float)
    cat_inc = defaultdict(float)
    for t in all_transactions:  # Показываем все категории включая личные
        if t["type"] == "expense": cat_exp[t["category"]] += t["debit"]
        else: cat_inc[t["category"]] += t["credit"]

    top_exp = sorted(cat_exp.items(), key=lambda x: -x[1])[:8]
    top_inc = sorted(cat_inc.items(), key=lambda x: -x[1])[:5]

    # По месяцам
    monthly = defaultdict(lambda: {"income": 0.0, "expense": 0.0})
    for t in business_txns:  # Только бизнес операции в динамике
        if t["type"] == "income": monthly[t["month"]]["income"] += t["credit"]
        else: monthly[t["month"]]["expense"] += t["debit"]
    monthly_sorted = [{"month": m, "income": clean(v["income"]), "expense": clean(v["expense"])}
                      for m, v in sorted(monthly.items())]

    # Категории которые никогда не бывают расходами
    INCOME_ONLY_CATS = {"Зарплата / доход", "Доход от клиентов", "Выручка Kaspi.kz", "Выручка по карточкам", "Торговая выручка", "Переводы входящие", "Выручка по карточкам Halyk"}

    # Расходы по категориям и месяцам
    months_list = [m["month"] for m in monthly_sorted]
    cat_monthly = defaultdict(lambda: defaultdict(float))
    for t in all_transactions:
        if t["type"] == "expense" and not t.get("internal") and t["category"] not in INCOME_ONLY_CATS:
            cat_monthly[t["category"]][t["month"]] += t["debit"]

    cat_monthly_table = []
    for cat in sorted(cat_monthly.keys()):
        row = {"category": cat, "total": clean(sum(cat_monthly[cat].values()))}
        for m in months_list:
            row[m] = clean(cat_monthly[cat].get(m, 0.0))
        row["share"] = round(row["total"] / total_exp * 100, 1) if total_exp > 0 else 0
        cat_monthly_table.append(row)
    cat_monthly_table.sort(key=lambda x: -x["total"])

    # Категории которые никогда не бывают поступлениями
    EXPENSE_ONLY_CATS = {"Налоги / взносы", "Аренда", "Транспорт", "Банковские расходы", "Кредит / займ", "Зарплата / ФОТ", "Реклама / маркетинг", "Процессинг / эквайринг", "Логистика / курьер",
                         "Связь / интернет", "Подписки / сервисы", "Образование",
                         "Здоровье / медицина", "Красота / уход", "IT / сервисы",
                         "Перевод собственнику", "Внутренний перевод"}

    # Поступления по категориям
    cat_income_data = defaultdict(float)
    for t in all_transactions:
        if t["type"] == "income" and not t.get("internal") and t["category"] not in EXPENSE_ONLY_CATS:
            cat_income_data[t["category"]] += t["credit"]
    cat_income_table = sorted(
        [{"category": cat, "total": clean(amt),
          "share": round(amt/total_inc*100, 1) if total_inc > 0 else 0}
         for cat, amt in cat_income_data.items()],
        key=lambda x: -x["total"]
    )

    # По банкам
    bank_data = defaultdict(lambda: {"income": 0.0, "expense": 0.0, "count": 0})
    for t in all_transactions:
        bank_data[t["bank"]]["income"]  += t["credit"]
        bank_data[t["bank"]]["expense"] += t["debit"]
        bank_data[t["bank"]]["count"]   += 1
    banks_list = [{"bank": b, "income": clean(v["income"]), "expense": clean(v["expense"]), "count": v["count"]}
                  for b, v in bank_data.items()]

    recommendations = generate_recommendations(all_transactions, total_inc, total_exp, net, cat_exp, monthly_sorted)

    return {
        "total": len(all_transactions),
        "total_income": total_inc,
        "total_expense": total_exp,
        "internal_out": internal_out,
        "internal_in": internal_in,
        "internal_count": len(internal_txns),
        "net_cashflow": net,
        "banks": banks_list,
        "months": months_list,
        "top_expenses": [{"category": k, "amount": clean(v)} for k,v in top_exp],
        "top_income":   [{"category": k, "amount": clean(v)} for k,v in top_inc],
        "monthly": monthly_sorted,
        "cat_monthly": cat_monthly_table,
        "cat_income": cat_income_table,
        "recommendations": recommendations,
    }

def generate_recommendations(txns, inc, exp, net, cat_exp, monthly):
    recs = []
    savings_rate = (net / inc * 100) if inc > 0 else 0
    if savings_rate < 10:
        recs.append({"type":"danger","icon":"alert","title":"Низкая норма сбережений",
            "text":f"Текущий уровень: {savings_rate:.1f}%. Рекомендуемый минимум — 20%. Проанализируйте топ расходных категорий."})
    elif savings_rate >= 30:
        recs.append({"type":"success","icon":"check","title":"Хорошая норма сбережений",
            "text":f"Уровень сбережений {savings_rate:.1f}% — выше нормы. Рассмотрите инвестиционные инструменты."})
    if net < 0:
        recs.append({"type":"danger","icon":"alert","title":"Отрицательный денежный поток",
            "text":f"Расходы превышают доходы на {abs(net):,.0f} ₸. Необходим срочный аудит бюджета."})
    top_cat = max(cat_exp.items(), key=lambda x: x[1]) if cat_exp else None
    if top_cat:
        share = top_cat[1] / exp * 100 if exp > 0 else 0
        if share > 30:
            recs.append({"type":"warning","icon":"info","title":f"Концентрация расходов: «{top_cat[0]}»",
                "text":f"{share:.1f}% всех расходов — одна категория ({top_cat[1]:,.0f} ₸). Диверсифицируйте бюджет."})
    if len(monthly) >= 3:
        cf = [m["income"]-m["expense"] for m in monthly[-3:]]
        if all(cf[i] < cf[i-1] for i in range(1,3)):
            recs.append({"type":"warning","icon":"trend","title":"Нисходящий тренд CF",
                "text":"Чистый CF снижается 3 месяца подряд. Проверьте источники доходов."})
        elif all(cf[i] > cf[i-1] for i in range(1,3)):
            recs.append({"type":"success","icon":"trend","title":"Положительный тренд",
                "text":"Денежный поток растёт 3 месяца подряд. Хороший момент для резервного фонда."})
    tax = cat_exp.get("Налоги / взносы", 0)
    if tax > 0:
        recs.append({"type":"info","icon":"info","title":"Налоговая нагрузка",
            "text":f"Уплачено налогов: {tax:,.0f} ₸ ({(tax/exp*100) if exp > 0 else 0:.1f}% расходов). Проверьте применение вычетов."})
    if not recs:
        recs.append({"type":"success","icon":"check","title":"Финансовые показатели в норме",
            "text":"Структура доходов и расходов сбалансирована."})
    return recs

# ─── РОУТЫ ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_file("index.html")

@app.route("/login", methods=["POST"])
def login():
    data = request.get_json()
    if data.get("password") == ACCESS_PASSWORD:
        session["auth"] = True
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "Неверный пароль"}), 401

@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})

def require_auth():
    return session.get("auth") is True

@app.route("/analyze", methods=["POST"])
def analyze():
    files = request.files.getlist("files")
    mode  = request.form.get("mode", "append")  # append | replace

    if not files:
        return jsonify({"error": "Файлы не загружены"}), 400

    upload_dir = "/tmp/busan_uploads"
    os.makedirs(upload_dir, exist_ok=True)

    new_transactions = []
    file_results = []
    for f in files:
        path = os.path.join(upload_dir, f.filename)
        f.save(path)
        bank = detect_bank(f.filename, path)
        txns = parse_file(path, bank)
        new_transactions.extend(txns)
        file_results.append({"name": f.filename, "bank": bank, "count": len(txns)})
        os.remove(path)

    if not new_transactions:
        return jsonify({"error": "Транзакции не найдены. Проверьте формат файлов."}), 400

    # Накопительная база
    db = load_db()
    if mode == "replace":
        db["transactions"] = new_transactions
        added = len(new_transactions)
    else:
        added = merge_transactions(db["transactions"], new_transactions)

    db["updated_at"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    save_db(db)

    all_transactions = db["transactions"]
    analytics = build_analytics(all_transactions)
    analytics["files"] = file_results
    analytics["added"] = added
    analytics["db_total"] = len(all_transactions)
    analytics["db_updated"] = db["updated_at"]
    analytics["mode"] = mode

    # Для Excel
    with open("/tmp/busan_txns.json", "w", encoding="utf-8") as f:
        json.dump(all_transactions, f, ensure_ascii=False)

    return jsonify(analytics)

@app.route("/db-stats", methods=["GET"])
def db_stats():
    db = load_db()
    txns = db["transactions"]
    return jsonify({
        "total": len(txns),
        "updated_at": db.get("updated_at"),
        "months": sorted(set(t["month"] for t in txns)),
        "banks": list(set(t["bank"] for t in txns)),
    })

@app.route("/db-clear", methods=["POST"])
def db_clear():
    save_db({"transactions": [], "updated_at": None})
    return jsonify({"ok": True})

@app.route("/db-analyze", methods=["GET"])
def db_analyze():
    """Показать аналитику по всей накопленной базе без загрузки файлов"""
    db = load_db()
    if not db["transactions"]:
        return jsonify({"error": "База данных пуста"}), 400
    analytics = build_analytics(db["transactions"])
    analytics["files"] = []
    analytics["db_total"] = len(db["transactions"])
    analytics["db_updated"] = db.get("updated_at")
    with open("/tmp/busan_txns.json", "w", encoding="utf-8") as f:
        json.dump(db["transactions"], f, ensure_ascii=False)
    return jsonify(analytics)

@app.route("/download-excel")
def download_excel():
    try:
        with open("/tmp/busan_txns.json", "r", encoding="utf-8") as f:
            txns = json.load(f)
        buf = generate_excel(txns)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        return send_file(buf, as_attachment=True,
                         download_name=f"BUSAN_Cashflow_{ts}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def generate_excel(transactions):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Дашборд"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    c = ws["A1"]; c.value = "BUSAN FINANCE  |  CASHFLOW ОТЧЁТ"
    c.fill = fill(NAVY); c.font = Font(name="Arial", bold=True, color=GOLD, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    ws.merge_cells("A2:G2")
    c = ws["A2"]; c.value = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    c.fill = fill(NAVY); c.font = font(color=GOLD, size=9, italic=True)
    c.alignment = Alignment(horizontal="center")
    total_inc = sum(t["credit"] for t in transactions)
    total_exp = sum(t["debit"]  for t in transactions)
    net = total_inc - total_exp
    kpis = [("Доходы (₸)", f"{total_inc:,.0f}"), ("Расходы (₸)", f"{total_exp:,.0f}"),
            ("Чистый CF (₸)", f"{net:+,.0f}"), ("Транзакций", str(len(transactions))),
            ("Банков", str(len(set(t['bank'] for t in transactions))))]
    for ci, (label, val) in enumerate(kpis, 1):
        c = ws.cell(row=4, column=ci, value=label)
        c.fill = fill(GOLD); c.font = font(bold=True, color=NAVY, size=9); c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=5, column=ci, value=val)
        c.fill = fill(LIGHT); c.font = Font(name="Arial", bold=True, color=NAVY, size=12)
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = 20

    # Транзакции
    ws2 = wb.create_sheet("Транзакции")
    ws2.sheet_view.showGridLines = False
    for ci, h in enumerate(["Дата","Банк","Тип","Категория","Описание","Расход","Доход"], 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.fill = fill(NAVY); c.font = font(bold=True, color=GOLD, size=10)
        c.alignment = Alignment(horizontal="center"); c.border = border_thin()
    ws2.column_dimensions["A"].width=12; ws2.column_dimensions["B"].width=12
    ws2.column_dimensions["C"].width=10; ws2.column_dimensions["D"].width=22
    ws2.column_dimensions["E"].width=45; ws2.column_dimensions["F"].width=16; ws2.column_dimensions["G"].width=16
    for ri, t in enumerate(sorted(transactions, key=lambda x: x["date"], reverse=True), 2):
        bg = LIGHT if ri%2==0 else WHITE
        for ci, v in enumerate([t["date"],t["bank"],"Расход" if t["type"]=="expense" else "Доход",
                                  t["category"],t["description"][:80],
                                  t["debit"] if t["debit"]>0 else "",
                                  t["credit"] if t["credit"]>0 else ""], 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            c.fill = fill(bg); c.border = border_thin()
            if ci==3: c.font = font(color=RED if v=="Расход" else GREEN, size=9, bold=True)
            elif ci in [6,7]: c.font = font(color=RED if ci==6 else GREEN, size=9); c.number_format="#,##0"
            else: c.font = font(color=NAVY, size=9)
    ws2.auto_filter.ref = f"A1:G{len(transactions)+1}"; ws2.freeze_panes = "A2"

    # Категории по месяцам
    ws4 = wb.create_sheet("Расходы по месяцам")
    ws4.sheet_view.showGridLines = False
    months = sorted(set(t["month"] for t in transactions))
    cat_monthly = defaultdict(lambda: defaultdict(float))
    for t in transactions:
        if t["type"] == "expense": cat_monthly[t["category"]][t["month"]] += t["debit"]
    hdrs = ["Категория"] + months + ["ИТОГО", "Уд. вес %"]
    for ci, h in enumerate(hdrs, 1):
        c = ws4.cell(row=1, column=ci, value=h)
        c.fill = fill(NAVY); c.font = font(bold=True, color=GOLD, size=9)
        c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = border_thin()
        ws4.column_dimensions[get_column_letter(ci)].width = 14
    ws4.column_dimensions["A"].width = 28; ws4.row_dimensions[1].height = 30
    cats_sorted = sorted(cat_monthly.keys(), key=lambda c: -sum(cat_monthly[c].values()))
    grand_total = sum(sum(cat_monthly[c].values()) for c in cats_sorted)
    for ri, cat in enumerate(cats_sorted, 2):
        bg = LIGHT if ri%2==0 else WHITE
        total = sum(cat_monthly[cat].values())
        share = round(total/grand_total*100, 1) if grand_total > 0 else 0
        row_vals = [cat] + [clean(cat_monthly[cat].get(m,0)) for m in months] + [clean(total), share]
        for ci, v in enumerate(row_vals, 1):
            c = ws4.cell(row=ri, column=ci, value=v if v != 0 else "")
            c.fill = fill(bg); c.border = border_thin()
            if ci == 1: c.font = font(color=NAVY, size=9, bold=True)
            elif ci == len(hdrs): c.font = font(color=NAVY, size=9)
            else: c.font = font(color=RED, size=9); c.number_format = "#,##0"
    ws4.freeze_panes = "B2"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ─── СКЛАД ────────────────────────────────────────────────────────────────────

from stock import parse_stock_report, build_stock_analytics

@app.route("/stock-analyze", methods=["POST"])
def stock_analyze():
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "Файл не загружен"}), 400

    upload_dir = "/tmp/busan_uploads"
    os.makedirs(upload_dir, exist_ok=True)

    all_items = []
    file_results = []
    for f in files:
        path = os.path.join(upload_dir, f.filename)
        f.save(path)
        items = parse_stock_report(path)
        all_items.extend(items)
        file_results.append({"name": f.filename, "count": len(items)})
        os.remove(path)

    if not all_items:
        return jsonify({"error": "Товары не найдены. Проверьте формат файла."}), 400

    analytics = build_stock_analytics(all_items)
    analytics["files"] = file_results

    with open("/tmp/busan_stock.json", "w", encoding="utf-8") as f:
        json.dump(all_items, f, ensure_ascii=False, default=str)

    return jsonify(analytics)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
